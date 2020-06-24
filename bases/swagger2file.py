# -*- coding: utf-8 -*-
# @Time    : 2020/6/23 18:07
# @Author  : grassroadsZ
# @File    : swagger2file.py

"""根据swagger 的api接口数据生成基础的业务及用例模块"""

import base64
import os

import jsonpath
from openpyxl import workbook

from bases.handle_requests import HandleRequest
from settings import (BUSINESS_DIR, CASE_DIR, DATA_DIR)


class SwaggerToFile(object):
    """swagger to file class"""

    def __init__(self, project_name: str, url: str, user=None, pwd=None):
        self.project_name = project_name
        self.swagger_url = url
        self.user = user
        self.pwd = pwd

    def open_swagger(self):
        """访问swagger"""
        # 存在需要登录认证的情况
        if self.user and self.pwd:
            s = f"{self.user}:{self.pwd}"
            auth_str = base64.b64encode(s.encode()).decode('utf-8')
            header = {"Authorization": f"Basic {auth_str}"}
            setattr(HandleRequest, "header", header)
            res = HandleRequest().send_request(method="get", path=self.swagger_url, verify=False)
        else:
            res = HandleRequest().send_request(method="get", path=self.swagger_url, verify=False)

        return res.json()

    def get_tags(self):
        """获取tags,根据tags进行模块分组"""

        tag_lists = [key.get("description").split() for key in t.open_swagger().get("tags")]
        tag_lists = ["".join(tag[:len(tag) - 1]) for tag in tag_lists]

        business_project_dir = os.path.join(BUSINESS_DIR, self.project_name)
        if not os.path.exists(business_project_dir):
            os.makedirs(business_project_dir)

        case_project_dir = os.path.join(CASE_DIR, self.project_name)

        if not os.path.exists(case_project_dir):
            os.makedirs(case_project_dir)

        for tag in tag_lists:
            file = "".join(list(map(lambda x: x.lower(), tag[:len(tag)])))

            if not os.path.exists(file):
                with open(os.path.join(business_project_dir, f"{file}.py"), encoding="utf-8", mode="a") as f:
                    f.write("# -.- encoding:utf-8")

            if not os.path.exists(file):
                with open(os.path.join(case_project_dir, f"test_{file}.py"), encoding="utf-8", mode="a") as f:
                    f.write("# -.- encoding:utf-8")

        return tag_lists

    def get_api_info(self):
        """
        '/bestSign/baiyong/certification/approve':{'post': {'tags': ['上上签接口'],
                                                'summary': '百用实名认证通过', 'operationId': 'approveBaiyongCertificationUsingPOST',
                                                 'consumes': ['application/json'], 'produces': ['*/*'],
                                                 'parameters': [{'name': 'TOKEN', 'in': 'query', 'description': 'TOKEN', 'required': True, 'type': 'string'}, {'in': 'body', 'name': 'baiyongApproveVO', 'description': 'baiyongApproveVO', 'required': True, 'schema': {'originalRef': 'BaiyongApproveVO', '$ref': '#/definitions/BaiyongApproveVO'}}],
                                                  'responses': {'200': {'description': 'OK', 'schema': {'type': 'string'}}, '201': {'description': 'Created'}, '401': {'description': 'Unauthorized'}, '403': {'description': 'Forbidden'}, '404': {'description': 'Not Found'}}, 'deprecated': False}}

        :param response: swagger接口响应的信息
        :return: ([[接口名称,请求路径,请求方法,请求头信息,请求参数],
                    [接口名称,请求路径,请求方法,请求头信息,请求参数]])
        """
        response = self.open_swagger().get('paths')
        result = []

        for key in response:

            _path = f"{key}"

            for method in response[key]:

                api_name = jsonpath.jsonpath(response[key][method], "$..summary")
                _api_name = api_name[0]

                content_type = jsonpath.jsonpath(response[key][method], "$..consumes")
                _content_type = content_type[0][0] if content_type else None
                header = {"Content-Type": _content_type}
                header["X-Access-Token"] = "token"

                # 取参数集合
                parameters = jsonpath.jsonpath(response[key][method], "$..parameters")
                # 最终需要发送数据的集合
                _data = {}
                # 存在parameters取不到的情况
                if not parameters:
                    pass
                else:
                    for param in parameters[0]:
                        # 取参数中key=name的值，将参数名称加到最终请求需要发送的数据集合,值默认为None
                        _data[param['name']] = None

                api_info = [_api_name, _path, method, header, _data]

            result.append(api_info)

        return result

    def generate_excel(self):
        wb = workbook.Workbook()
        sheet_header = ["case_id", "case_name", "path", "method", "headers", "data"]

        api_info = self.get_api_info()
        modules = self.get_tags()
        for sheet in modules:
            # 表单sheet页名
            ws = wb.create_sheet(title=sheet, index=-1)
            for index, name in enumerate(sheet_header):
                # 写入表头
                ws.cell(1, index + 1).value = name

            row = 2
            for api in api_info:
                # 写入内容

                for index, value in enumerate(api):
                    ws.cell(index + 2, 1).value = index + 1
                    ws.cell(row, index + 2).value = str(value)
                row += 1

        file = os.path.join(DATA_DIR, f"{self.project_name}_api.xlsx")
        setting_file = os.path.join(DATA_DIR, f"{self.project_name}_settings.py")
        if not os.path.exists(setting_file):
            string = "# -.- encoding:utf-8"
            with open(setting_file, encoding="utf-8", mode='a') as f:
                f.write(string)

        if not os.path.exists(file):
            wb.save(file)
        else:
            print(f"{file}以存在")


if __name__ == '__main__':
    t = SwaggerToFile("accm", url="http://baidu.com.cn/api-docs", user=None, pwd=None)
    print(t.generate_excel())
