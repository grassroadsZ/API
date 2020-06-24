'''
-*-conding:utf-8
@Time:2019-05-21 6:40
@auther:grassroadsZ
@file:handle_excel.py
'''
from openpyxl import load_workbook
from collections import namedtuple


class ExcelOption(object):
    """
    excel 操作类
    """

    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active
        self.all_sheet = self.wb.sheetnames

    def get_cases(self, sheet_name):
        """
        获取所有的测试用例
        :return:所有的测试用例
        """

        self.ws = self.wb[sheet_name]
        self.case_list = []
        self.Case = namedtuple(
            "CaseData", tuple(self.ws.iter_rows(max_row=1, values_only=True))[0])
        Case = tuple(self.ws.iter_rows(min_row=2, values_only=True))
        for data in Case:
            self.case_list.append(self.Case(*data))
        return self.case_list

    def get_all_sheet(self):
        return self.all_sheet[:len(self.all_sheet) - 1]


if __name__ == '__main__':
    excel = ExcelOption("api.xlsx")
    for sheet in excel.all_sheet:
        print(excel.get_cases(sheet))
